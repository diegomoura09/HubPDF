"""
PDF tools routes
"""
import uuid
import io
from typing import List
from fastapi import APIRouter, Request, Depends, UploadFile, File, Form, HTTPException, status, BackgroundTasks
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.auth import require_auth
from app.models import User
from app.services.pdf_service import PDFService
from app.services.quota_service import QuotaService
from app.services.file_service import FileService
from app.services.i18n import get_user_locale, get_translations, translate
from app.template_helpers import templates
from app.utils.validators import FileValidator

router = APIRouter()
pdf_service = PDFService()
quota_service = QuotaService()
file_service = FileService()

def cleanup_job_files(user_id: int, job_id: str):
    """Background task to cleanup job files after delay"""
    import time
    import threading
    
    def delayed_cleanup():
        time.sleep(1800)  # 30 minutes
        file_service.cleanup_job_directory(user_id, job_id)
    
    # Use threading instead of asyncio to avoid event loop issues
    cleanup_thread = threading.Thread(target=delayed_cleanup, daemon=True)
    cleanup_thread.start()

@router.get("/", response_class=HTMLResponse)
async def tools_index(
    request: Request,
    user: User = Depends(require_auth)
):
    """Tools overview page"""
    locale = get_user_locale(request)
    translations = get_translations(locale)
    
    return templates.TemplateResponse(
        "tools/index.html",
        {
            "request": request,
            "user": user,
            "locale": locale,
            "translations": translations
        }
    )

@router.get("/merge", response_class=HTMLResponse)
async def merge_page(
    request: Request,
    user: User = Depends(require_auth)
):
    """PDF merge tool page"""
    locale = get_user_locale(request)
    translations = get_translations(locale)
    
    return templates.TemplateResponse(
        "tools/merge.html",
        {
            "request": request,
            "user": user,
            "locale": locale,
            "translations": translations
        }
    )

@router.post("/merge")
async def merge_pdfs(
    request: Request,
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Merge multiple PDF files"""
    try:
        locale = get_user_locale(request)
        
        # Check quota
        allowed, message, requires_watermark = quota_service.check_operation_allowed(db, user)
        if not allowed:
            raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail=message)
        
        # Validate files
        FileValidator.validate_multiple_files(files, max_count=10)
        
        for file in files:
            # Check file size
            file_allowed, size_message = quota_service.check_file_size_allowed(user, file.size)
            if not file_allowed:
                raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail=size_message)
            
            # Validate PDF
            FileValidator.validate_pdf_file(file, quota_service.get_plan_limits(user.plan)["max_file_size"])
        
        # Create job directory
        job_id = file_service.create_job_directory(user.id)
        
        # Process files
        pdf_files = []
        for file in files:
            file.file.seek(0)
            pdf_files.append(file.file)
        
        # Get watermark text if needed
        watermark_text = ""
        if requires_watermark:
            watermark_text = translate("watermark_text", locale)
        
        # Merge PDFs
        result_bytes = pdf_service.merge_pdfs(
            pdf_files, user.id, job_id, requires_watermark, watermark_text
        )
        
        # Save result
        result_filename = "merged.pdf"
        file_service.save_result_file(user.id, job_id, result_bytes, result_filename)
        
        # Increment quota
        quota_service.increment_operation_count(db, user)
        
        # Schedule cleanup
        background_tasks.add_task(cleanup_job_files, user.id, job_id)
        
        # Return download response
        return StreamingResponse(
            iter([result_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=merged.pdf"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to merge PDFs: {str(e)}"
        )

@router.get("/split", response_class=HTMLResponse)
async def split_page(
    request: Request,
    user: User = Depends(require_auth)
):
    """PDF split tool page"""
    locale = get_user_locale(request)
    translations = get_translations(locale)
    
    return templates.TemplateResponse(
        "tools/split.html",
        {
            "request": request,
            "user": user,
            "locale": locale,
            "translations": translations
        }
    )

@router.post("/split")
async def split_pdf(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    page_ranges: str = Form(...),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Split PDF by page ranges"""
    try:
        locale = get_user_locale(request)
        
        # Check quota
        allowed, message, requires_watermark = quota_service.check_operation_allowed(db, user)
        if not allowed:
            raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail=message)
        
        # Validate file
        file_allowed, size_message = quota_service.check_file_size_allowed(user, file.size)
        if not file_allowed:
            raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail=size_message)
        
        FileValidator.validate_pdf_file(file, quota_service.get_plan_limits(user.plan)["max_file_size"])
        
        # Create job directory
        job_id = file_service.create_job_directory(user.id)
        
        # Get watermark text if needed
        watermark_text = ""
        if requires_watermark:
            watermark_text = translate("watermark_text", locale)
        
        # Split PDF
        result_bytes = pdf_service.split_pdf(
            file.file, page_ranges, user.id, job_id, requires_watermark, watermark_text
        )
        
        # Increment quota
        quota_service.increment_operation_count(db, user)
        
        # Schedule cleanup
        background_tasks.add_task(cleanup_job_files, user.id, job_id)
        
        # Return download response
        return StreamingResponse(
            iter([result_bytes]),
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=split_pdfs.zip"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to split PDF: {str(e)}"
        )

@router.get("/compress", response_class=HTMLResponse)
async def compress_page(
    request: Request,
    user: User = Depends(require_auth)
):
    """PDF compress tool page"""
    locale = get_user_locale(request)
    translations = get_translations(locale)
    
    return templates.TemplateResponse(
        "tools/compress.html",
        {
            "request": request,
            "user": user,
            "locale": locale,
            "translations": translations
        }
    )

@router.post("/compress")
async def compress_pdf(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Compress PDF file"""
    try:
        locale = get_user_locale(request)
        
        # Check quota
        allowed, message, requires_watermark = quota_service.check_operation_allowed(db, user)
        if not allowed:
            raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail=message)
        
        # Validate file
        file_allowed, size_message = quota_service.check_file_size_allowed(user, file.size)
        if not file_allowed:
            raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail=size_message)
        
        FileValidator.validate_pdf_file(file, quota_service.get_plan_limits(user.plan)["max_file_size"])
        
        # Create job directory
        job_id = file_service.create_job_directory(user.id)
        
        # Get watermark text if needed
        watermark_text = ""
        if requires_watermark:
            watermark_text = translate("watermark_text", locale)
        
        # Compress PDF
        result_bytes = pdf_service.compress_pdf(
            file.file, user.id, job_id, requires_watermark, watermark_text
        )
        
        # Increment quota
        quota_service.increment_operation_count(db, user)
        
        # Schedule cleanup
        background_tasks.add_task(cleanup_job_files, user.id, job_id)
        
        # Return download response
        return StreamingResponse(
            iter([result_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=compressed.pdf"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to compress PDF: {str(e)}"
        )

@router.get("/pdf-to-images", response_class=HTMLResponse)
async def pdf_to_images_page(
    request: Request,
    user: User = Depends(require_auth)
):
    """PDF to images tool page"""
    locale = get_user_locale(request)
    translations = get_translations(locale)
    
    return templates.TemplateResponse(
        "tools/pdf_to_images.html",
        {
            "request": request,
            "user": user,
            "locale": locale,
            "translations": translations
        }
    )

@router.post("/pdf-to-images")
async def pdf_to_images(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    format: str = Form(default="png"),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Convert PDF to images"""
    try:
        locale = get_user_locale(request)
        
        # Check quota
        allowed, message, requires_watermark = quota_service.check_operation_allowed(db, user)
        if not allowed:
            raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail=message)
        
        # Validate format
        if format not in ["png", "jpg", "jpeg"]:
            format = "png"
        
        # Validate file
        file_allowed, size_message = quota_service.check_file_size_allowed(user, file.size)
        if not file_allowed:
            raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail=size_message)
        
        FileValidator.validate_pdf_file(file, quota_service.get_plan_limits(user.plan)["max_file_size"])
        
        # Create job directory
        job_id = file_service.create_job_directory(user.id)
        
        # Get watermark text if needed
        watermark_text = ""
        if requires_watermark:
            watermark_text = translate("watermark_text", locale)
        
        # Convert PDF to images
        result_bytes = pdf_service.pdf_to_images(
            file.file, user.id, job_id, format, requires_watermark, watermark_text
        )
        
        # Increment quota
        quota_service.increment_operation_count(db, user)
        
        # Schedule cleanup
        background_tasks.add_task(cleanup_job_files, user.id, job_id)
        
        # Return download response
        return StreamingResponse(
            iter([result_bytes]),
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename=pdf_images.zip"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to convert PDF to images: {str(e)}"
        )

@router.get("/images-to-pdf", response_class=HTMLResponse)
async def images_to_pdf_page(
    request: Request,
    user: User = Depends(require_auth)
):
    """Images to PDF tool page"""
    locale = get_user_locale(request)
    translations = get_translations(locale)
    
    return templates.TemplateResponse(
        "tools/images_to_pdf.html",
        {
            "request": request,
            "user": user,
            "locale": locale,
            "translations": translations
        }
    )

@router.post("/images-to-pdf")
async def images_to_pdf(
    request: Request,
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Convert images to PDF"""
    try:
        locale = get_user_locale(request)
        
        # Check quota
        allowed, message, requires_watermark = quota_service.check_operation_allowed(db, user)
        if not allowed:
            raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail=message)
        
        # Validate files
        FileValidator.validate_multiple_files(files, max_count=20)
        
        for file in files:
            file_allowed, size_message = quota_service.check_file_size_allowed(user, file.size)
            if not file_allowed:
                raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail=size_message)
            
            FileValidator.validate_image_file(file, quota_service.get_plan_limits(user.plan)["max_file_size"])
        
        # Create job directory
        job_id = file_service.create_job_directory(user.id)
        
        # Process files
        image_files = []
        for file in files:
            file.file.seek(0)
            image_files.append(file.file)
        
        # Get watermark text if needed
        watermark_text = ""
        if requires_watermark:
            watermark_text = translate("watermark_text", locale)
        
        # Convert images to PDF
        result_bytes = pdf_service.images_to_pdf(
            image_files, user.id, job_id, requires_watermark, watermark_text
        )
        
        # Increment quota
        quota_service.increment_operation_count(db, user)
        
        # Schedule cleanup
        background_tasks.add_task(cleanup_job_files, user.id, job_id)
        
        # Return download response
        return StreamingResponse(
            iter([result_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=images.pdf"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to convert images to PDF: {str(e)}"
        )

@router.get("/extract-text", response_class=HTMLResponse)
async def extract_text_page(
    request: Request,
    user: User = Depends(require_auth)
):
    """Extract text tool page"""
    locale = get_user_locale(request)
    translations = get_translations(locale)
    
    return templates.TemplateResponse(
        "tools/extract_text.html",
        {
            "request": request,
            "user": user,
            "locale": locale,
            "translations": translations
        }
    )

@router.post("/extract-text")
async def extract_text(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    output_format: str = Form(default="display"),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Extract text from PDF"""
    try:
        locale = get_user_locale(request)
        
        # Check quota
        allowed, message, requires_watermark = quota_service.check_operation_allowed(db, user)
        if not allowed:
            raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail=message)
        
        # Validate file
        file_allowed, size_message = quota_service.check_file_size_allowed(user, file.size)
        if not file_allowed:
            raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail=size_message)
        
        FileValidator.validate_pdf_file(file, quota_service.get_plan_limits(user.plan)["max_file_size"])
        
        # Create job directory
        job_id = file_service.create_job_directory(user.id)
        
        # Extract text
        extracted_text = pdf_service.extract_text(file.file, user.id, job_id)
        
        # Increment quota
        quota_service.increment_operation_count(db, user)
        
        # Schedule cleanup
        background_tasks.add_task(cleanup_job_files, user.id, job_id)
        
        # Handle different output formats
        if output_format == "display":
            # Return JSON response for HTMX to display
            return JSONResponse({
                "success": True,
                "text": extracted_text,
                "message": "Texto extraído com sucesso"
            })
        elif output_format == "txt":
            # Return text as TXT download
            return StreamingResponse(
                iter([extracted_text.encode('utf-8')]),
                media_type="text/plain",
                headers={"Content-Disposition": "attachment; filename=extracted_text.txt"}
            )
        elif output_format == "xlsx":
            # Create Excel file with extracted text
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Extracted Text"
            
            # Add header
            ws['A1'] = "Texto Extraído do PDF"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Add extracted text (split by lines for better formatting)
            lines = extracted_text.split('\n')
            for i, line in enumerate(lines, start=3):  # Start from row 3
                ws[f'A{i}'] = line
                ws[f'A{i}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            # Adjust column width
            ws.column_dimensions['A'].width = 100
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return StreamingResponse(
                iter([excel_buffer.read()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=extracted_text.xlsx"}
            )
        else:
            # Default to display
            return JSONResponse({
                "success": True,
                "text": extracted_text,
                "message": "Texto extraído com sucesso"
            })
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to extract text: {str(e)}"
        )
